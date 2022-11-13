from openpyxl import Workbook
from openpyxl.styles import Font, colors, Border, Side, Alignment, PatternFill
from openpyxl.chart import Reference, Series, ScatterChart

print("================================")
print("   SEGTs IOU EXCEL GENERATOR")
print("================================")

WorkSheetName = input("Enter the name of the dataset: ")
DevIOUMean = input("Enter Dev IOU Mean Array: ")
NumberOfClasses = input("How many classes does the dataset have: ")
LogAdder = input("What is the LogAdder value: ")
LogMultiplier = input("What is the LogMultiplier value: ")
TolRange = input("What is the ToleranceRange value: ")

Array = DevIOUMean.split(";")
Array = [float(Idx) for Idx in Array]

wb = Workbook()

ws1 = wb.active
ws1.title = WorkSheetName

br = Border(left = Side(border_style = 'thin'), right = Side(border_style = 'thin'), top = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))
al = Alignment(horizontal = 'center')
pf = PatternFill(fill_type = 'lightGrid', start_color = "00FFFF99", end_color = "00FFFF99")

ws1['A1'] = "EPOCH"
A1 = ws1['A1']
A1.border = br
A1.alignment = al
A1.fill = pf

for i in range(len(Array)):
    ws1.cell(column = 1, row = i+2, value = i+1)

ws1['B1'] = "IOU"
B1 = ws1['B1']
B1.border = br
B1.alignment = al
B1.fill = pf

for j in range(len(Array)):
    ws1.cell(column = 2, row = j+2, value = Array[j])
  
ws1['C1'] = "LOG UP"
C1 = ws1['C1']
C1.border = br
C1.alignment = al
C1.fill = pf

ws1['D1'] = "LOG LOW"
D1 = ws1['D1']
D1.border = br
D1.alignment = al
D1.fill = pf

for n in range(len(Array)):
    ws1.cell(column = 3, row = n+2, value = '=D'+str(n+2)+'+$H$4')
    ws1.cell(column = 4, row = n+2, value = '=LOG((A'+str(n+2)+'+$H$2),10)*$H$3+((100/$H$5)-5)')

for m in range(5):
 ws1.merge_cells('F'+str(m+1)+':G'+str(m+1))

ws1['F1'] = "BEST IOU MEAN"
ws1['F2'] = "LOG ADDER"
ws1['F3'] = "LOG MULTIPLIER"
ws1['F4'] = "TOLERANCE RANGE"
ws1['F5'] = "CLASSES"
ws1['H1'] = '=MAX(B2:B'+str(len(Array) + 1)+')'
ws1['H2'] = LogAdder
ws1['H3'] = LogMultiplier
ws1['H4'] = TolRange
ws1['H5'] = NumberOfClasses

for k in range(2, len(Array) + 2):
    A = ws1['A' + str(k)]
    A.border = br
    A.alignment = al
    B = ws1['B' + str(k)]
    B.border = br
    B.alignment = al
    C = ws1['C' + str(k)]
    C.border = br
    C.alignment = al
    D = ws1['D' + str(k)]
    D.border = br
    D.alignment = al

for l in range(5):
    F = ws1['F' + str(l+1)]
    F.border = br
    F.fill = pf
    F.alignment = al
    G = ws1['G' + str(l+1)]
    G.border = br
    G.fill = pf
    H = ws1['H' + str(l+1)]
    H.border = br
    H.alignment = al

#Drawing Graph
Chart = ScatterChart()
Chart.title = "IOU VS EPOCH GRAPH"
Chart.x_axis.title = "EPOCH"
Chart.y_axis.title = "IOU"

XValues = Reference(ws1, min_col = 1, min_row = 2, max_row = int(len(Array)) + 1)
for p in range(2, 5):
    YValues = Reference(ws1, min_col = p, min_row = 1, max_row = int(len(Array)) + 1)
    CSeries = Series(YValues, XValues, title_from_data = True)
    Chart.series.append(CSeries)

ws1.add_chart(Chart, "F8")

wb.save(WorkSheetName + "IOUTolReport.xlsx")


